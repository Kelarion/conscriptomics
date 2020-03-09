"""
Here we attempt to produce a nice schedule for the CTN lab meetings. The goal 
is to automatically keep track of who is eligible to present, and then order 
them. It would be good to include support for hard conflicts (e.g. 'plan to 
leave the lab by this date'), or even pairing preferences.

This function gives n samples from the available speakers, taken without 
replacement. 

You must have a list of current members saved somewhere, as an *.xlsx file.

DEPENDENCIES:
Ideally, this is all in base python. That might not be possible, so 
dependencies should go here:
    - openpyxl (https://openpyxl.readthedocs.io/en/stable/index.html)

"""
#%%
from openpyxl import Workbook, load_workbook
from datetime import datetime
import string
import csv
import random
import math
import re

letters = string.ascii_lowercase
today = datetime.today().year + (datetime.today().month-1)/12
#%% helper functions
def recency_weights(lastpres, today):
    '''
    return weights, which increase as a function of today-lastpres
    
    right now, weights are today-lastpres passed through a sigmoid centred at
    0.5 (i.e. 6 months) and with a slope of 10.
    '''
    dt = [min([today - l, 10]) for l in lastpres]
    
    link = lambda x:math.exp(10*x-10)/(math.exp(10*x-10) + 1)
    w = [link(d) for d in dt]
    
    return w

def weighted_shuffle(L, weights):
    '''
    shuffle L, but with weights skewing the order
    '''
    x = [None for _ in L]
    for i in range(len(L)): # trick for weighted sampling w/out replacement
        q = random.choices(L, weights)[0]
        weights[L.index(q)] = 0
        x[i] = q
    
    return x

def robust_find(query, names):
    '''
    Returns which of 'names' is present in query
    '''
    
    reg = '(\w+(?:\s|\w+\.|\(\w+\))*\w+)' # for finding names
    surnames = [re.findall('(\w+)',n)[-1] for n in names]
    givnames = [re.findall('(\w+)',n)[0] for n in names]
    
    return [sum([(n in r) and (m in r) for r in re.findall(reg, query)])>0
             for n,m in zip(surnames,givnames)]

#%%
svdir = '/home/matteo/Documents/uni/columbia/duties/' # where to keep the spreadsheet
nslots = 26 # how many slots do we need to fill?

# in a misguided attempt to minimise dependencies, there will be
# a lot of ugly list comprehensions ahead, sorry in advance.

# ----------------------------------------------------------------
# READ THE LIST OF CURRENT MEMBERS
# load *.xlsx file
wb = load_workbook(filename=svdir+'ctn_members_ay20.xlsx')

ws = wb.active
headers = ws['1']
headers = [h.value.lower() for h in headers]

is_uni = ['uni' in h for h in headers].index(True) # find the column indices
is_name1 = ['first' in h for h in headers].index(True)
is_name2 = ['last' in h for h in headers].index(True)
#is_pres = ['presented' in h for h in headers].index(True)
is_cons = ['consider' in h for h in headers].index(True)
is_startdate = ['start' in h for h in headers].index(True)

unis = [w.value for w in ws[letters[is_uni]][1:]] # get the values
names = [w1.value + ' ' + w2.value \
         for w1,w2 in zip(ws[letters[is_name1]][1:],ws[letters[is_name2]][1:])]
#surnames = [w.value for w in ws[letters[is_name2]][1:]]
#givnames = [w.value for w in ws[letters[is_name1]][1:]]
#presented = [w.value for w in ws[letters[is_pres]][1:]]
consider = [w.value for w in ws[letters[is_cons]][1:]]
started = [w.value if type(w.value) is datetime else datetime.today() \
           for w in ws[letters[is_startdate]][1:]]

# here we define 'eligible': able to speak (i.e. postdoc or y3+ grad student)
# and having arrived at least 2 months ago
cons = [c is not None and (d.year+d.month/12 < today-(1/6)) for c,d in zip(consider,started)]

# ----------------------------------------------------------------
# READ THE ARCHIVE
with open(svdir+'archive.csv', 'r') as f:
    reader = csv.DictReader(f)
    last_pres = [None for _ in range(len(unis))]
    for r in reader:
        foo = robust_find(r['Speaker'], names) # defined above
        if any(foo):
            inds = [i for i,x in enumerate(foo) if x]
            for i in inds:
                if last_pres[i] is None:
                    d = datetime.strptime(r['Date '], '%m/%d/%Y')
                    last_pres[i] = round(d.year + d.month/12, 3)

pool = cons
with open(svdir+'list.csv', 'w', newline='') as f:
    writer = csv.writer(f, delimiter=',',
                        quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['name','uni','eligibility','last_presentation'])
    for i, vals in enumerate(zip(names, unis, pool, last_pres)):
        y = vals[3] if vals[3] is not None else 1.0
        writer.writerow([vals[0], vals[1], vals[2], y])    

# ----------------------------------------------------------------
# LOAD SPEAKER INFORMATION
with open(svdir+'list.csv', 'r') as f:
    reader = csv.DictReader(f)
    pool = [(r['eligibility']=='True') for r in reader]
    f.seek(0)
    reader = csv.DictReader(f)
    lastpres = [float(r['last_presentation']) for r in reader]
    
notrecent = [l <= today-0.3 for l in lastpres]
pool = [p and n for p,n in zip(pool, notrecent)]

if sum(pool) == 0:
    print('Everyone has presented! Yay! Time to reset...')
    pool = cons

# ----------------------------------------------------------------
# ORDER THE SPEAKERS
# the idea here is that when we make a schedule, we subsample from the set of 
# all potential speakers at that time. but when we make the next schedule, 
# there will be new people potentially. we want to ensure that we've made a 
# full pass through the original set of speakers before we include those new
# people on the cue -- no one escapes their duty. furthermore

# shuffle the eligible pool
idx = random.sample(list(range(len(unis))), k=len(unis))
idx = [i for i in idx if pool[i]]

# this is a little weird: we would like the new order to respect how 
# recently everyone spoke. rather than enforce a hard limit, we will try
# weighting people's position in the queue based on recency. maybe having
# a fixed order would be better, if less clever.

w = recency_weights([l for i,l in enumerate(lastpres) if pool[i]], today)
idx = weighted_shuffle(idx, w) # custom functions, defined above

# deal with the remainders
if sum(pool) < nslots: 
    # in this scenario: we've used up the pool of speakers established in the
    # last scheduling. so, we'll add any eligible newcomers to the queue and, 
    # after that, reset the pool to include all currently eligible speakers.
    
    df = nslots-sum(pool)
    
    nidx = random.sample(list(range(len(unis))), k=len(unis)) # first get new people
    nidx = [i for i in nidx if cons[i] and (i not in idx)]
    
    allidx = [i for i in range(len(unis)) if cons[i] and (i not in idx) and (i not in nidx)]
    
    w = recency_weights([l for i,l in enumerate(lastpres) if cons[i] \
                         and (i not in idx) and (i not in nidx)], today)
    nidx += weighted_shuffle(allidx, w) # custom functions, defined above
    
    idx += nidx[:df]
    
    # for next time: we'll make the schedule based on everyone who's currently
    # eligible, but hasn't just been assigned to speak
    new_pool = [c and (i not in idx) for i,c in enumerate(cons)]
else: 
    # this is easier: just subsample and make note of who was assigned to speak
    idx = idx[:nslots]
    
    new_pool = [p and not (i in idx) for i,p in enumerate(pool)]

# ----------------------------------------------------------------
# SAVE ORDER
# write the new schedule into a csv file
with open(svdir+'schedule.csv', 'w', newline='') as outf:
    writer = csv.writer(outf, delimiter=',',
                     quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['name','uni'])
    for i in idx:
        writer.writerow([names[i], unis[i]])


    
    